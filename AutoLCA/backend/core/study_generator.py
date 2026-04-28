"""
ISO 14044 LCA Study Package Generator
Produces a ZIP containing all documents required for a recognized LCA study:
  01_Goal_and_Scope.pdf
  02_Life_Cycle_Inventory.xlsx
  03_LCIA_Results.xlsx
  04_LCIA_Report.pdf  (extends JRCReporter)
  05_Interpretation.pdf
  06_Data_Quality_Assessment.pdf
  07_Critical_Review_Checklist.pdf
  metadata.json
"""
import os
import io
import json
import logging
import datetime
import zipfile
from typing import Dict, List, Any, Optional

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

logger = logging.getLogger(__name__)

# Standard JRC EF 3.1 unit map
UNIT_MAP = {
    'gwp_climate_change': 'kg CO2 eq',
    'odp_ozone_depletion': 'kg CFC11 eq',
    'ap_acidification': 'mol H+ eq',
    'ep_freshwater': 'kg P eq',
    'ep_marine': 'kg N eq',
    'ep_terrestrial': 'mol N eq',
    'pocp_photochemical_ozone': 'kg NMVOC eq',
    'pm_particulate_matter': 'disease inc.',
    'ir_ionising_radiation': 'kBq U235 eq',
    'ht_c_human_toxicity_cancer': 'CTUh',
    'ht_nc_human_toxicity_non_cancer': 'CTUh',
    'et_fw_ecotoxicity_freshwater': 'CTUe',
    'lu_land_use': 'Pt',
    'wsf_water_scarcity': 'm3 world eq',
    'ru_mm_resource_use_min_met': 'kg Sb eq',
    'ru_f_resource_use_fossils': 'MJ'
}

IMPACT_CATEGORIES = list(UNIT_MAP.keys())


class StudyPackageGenerator:
    """
    Generates all files needed for a recognized ISO 14040/14044 LCA study.
    """

    def __init__(self):
        self.styles = getSampleStyleSheet()
        self._define_custom_styles()

    def _define_custom_styles(self):
        self.header_style = ParagraphStyle(
            'StudyHeader', parent=self.styles['Normal'],
            fontSize=18, fontName='Helvetica-Bold',
            textColor=colors.toColor("#1a3a4a"),
            alignment=1, spaceAfter=24
        )
        self.section_style = ParagraphStyle(
            'StudySection', parent=self.styles['Heading2'],
            fontSize=13, fontName='Helvetica-Bold',
            textColor=colors.toColor("#2d5f7a"),
            spaceBefore=20, spaceAfter=12,
            backColor=colors.toColor("#f0f4f8"),
            borderPadding=6
        )
        self.body_style = ParagraphStyle(
            'StudyBody', parent=self.styles['Normal'],
            fontSize=10, leading=14, spaceAfter=8
        )
        self.small_style = ParagraphStyle(
            'StudySmall', parent=self.styles['Normal'],
            fontSize=8, textColor=colors.grey, spaceAfter=6
        )

    def generate_package(
        self,
        goal_and_scope: Dict[str, Any],
        nodes: List[Dict[str, Any]],
        edges: List[Dict[str, Any]],
        lcia_results: Dict[str, Any],
        validation_report: Optional[Dict[str, Any]] = None,
        sensitivity_results: Optional[Dict[str, Any]] = None,
        framework: str = "iso-14044"
    ) -> io.BytesIO:
        """
        Generates a ZIP buffer containing all study documents.
        """
        zip_buffer = io.BytesIO()

        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            # 01. Goal & Scope
            gs_pdf = self._generate_goal_and_scope_pdf(goal_and_scope, framework)
            zf.writestr("01_Goal_and_Scope.pdf", gs_pdf.getvalue())

            # 02. Life Cycle Inventory (Excel)
            lci_xlsx = self._generate_lci_excel(nodes, edges)
            zf.writestr("02_Life_Cycle_Inventory.xlsx", lci_xlsx.getvalue())

            # 03. LCIA Results (Excel)
            lcia_xlsx = self._generate_lcia_excel(lcia_results)
            zf.writestr("03_LCIA_Results.xlsx", lcia_xlsx.getvalue())

            # 04. LCIA Report (PDF)
            lcia_pdf = self._generate_lcia_report_pdf(lcia_results, goal_and_scope, framework)
            zf.writestr("04_LCIA_Report.pdf", lcia_pdf.getvalue())

            # 05. Interpretation Report
            interp_pdf = self._generate_interpretation_pdf(
                lcia_results, sensitivity_results, goal_and_scope
            )
            zf.writestr("05_Interpretation.pdf", interp_pdf.getvalue())

            # 06. Data Quality Assessment
            dq_pdf = self._generate_data_quality_pdf(nodes, goal_and_scope)
            zf.writestr("06_Data_Quality_Assessment.pdf", dq_pdf.getvalue())

            # 07. Critical Review Checklist
            checklist_pdf = self._generate_review_checklist_pdf(
                validation_report, goal_and_scope, framework
            )
            zf.writestr("07_Critical_Review_Checklist.pdf", checklist_pdf.getvalue())

            # metadata.json
            meta = self._generate_metadata(goal_and_scope, lcia_results, framework)
            zf.writestr("metadata.json", json.dumps(meta, indent=2))

        zip_buffer.seek(0)
        return zip_buffer

    # ═══════════════════════════════════════════════════════════════════════
    # 01. GOAL & SCOPE PDF
    # ═══════════════════════════════════════════════════════════════════════

    def _generate_goal_and_scope_pdf(self, gs: Dict, framework: str) -> io.BytesIO:
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=50, bottomMargin=50)
        elements = []

        elements.append(Paragraph("GOAL AND SCOPE DEFINITION", self.header_style))
        elements.append(Paragraph(
            f"<i>Compliant with {framework.upper().replace('-', ' ')} | "
            f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}</i>",
            self.small_style
        ))
        elements.append(Spacer(1, 12))

        # 1.1 Goal
        elements.append(Paragraph("1.1 Goal of the Study", self.section_style))
        elements.append(Paragraph(
            f"<b>Project Title:</b> {gs.get('projectTitle', 'Untitled')}",
            self.body_style
        ))
        elements.append(Paragraph(
            f"<b>Intended Application:</b> {gs.get('intendedApplication', 'Not specified')}",
            self.body_style
        ))
        elements.append(Paragraph(
            f"<b>Reasons for Carrying Out the Study:</b> {gs.get('reasons', 'Not stated')}",
            self.body_style
        ))
        elements.append(Paragraph(
            f"<b>Intended Audience:</b> {gs.get('intendedAudience', 'Not specified')}",
            self.body_style
        ))
        is_comp = gs.get("isComparativePublic", False)
        elements.append(Paragraph(
            f"<b>Comparative Assertion Intended for Public Disclosure:</b> {'Yes' if is_comp else 'No'}",
            self.body_style
        ))

        # 1.2 Scope - Functional Unit
        elements.append(Paragraph("1.2 Functional Unit", self.section_style))
        fu = gs.get("functionalUnit", {})
        if isinstance(fu, dict):
            elements.append(Paragraph(
                f"<b>Description:</b> {fu.get('description', 'Not defined')}",
                self.body_style
            ))
            elements.append(Paragraph(
                f"<b>Magnitude:</b> {fu.get('magnitude', 1.0)} {fu.get('unit', 'unit')}",
                self.body_style
            ))
            elements.append(Paragraph(
                f"<b>Reference Flow:</b> {fu.get('referenceFlow', 'Not specified')}",
                self.body_style
            ))
        else:
            elements.append(Paragraph(f"Functional Unit: {fu}", self.body_style))

        # 1.3 System Boundary
        elements.append(Paragraph("1.3 System Boundary", self.section_style))
        sb = gs.get("systemBoundary", {})
        if isinstance(sb, dict):
            elements.append(Paragraph(f"<b>Scope:</b> {sb.get('scope', 'Not defined')}", self.body_style))
            elements.append(Paragraph(f"<b>Capital Goods Included:</b> {'Yes' if sb.get('capitalGoods') else 'No'}", self.body_style))
            elements.append(Paragraph(f"<b>Packaging Included:</b> {'Yes' if sb.get('packaging') else 'No'}", self.body_style))
            elements.append(Paragraph(f"<b>Cut-off Threshold:</b> {sb.get('cutoffThreshold', 0)*100:.1f}%", self.body_style))
            excluded = sb.get("excludedFlows", [])
            if excluded:
                elements.append(Paragraph(f"<b>Excluded Flows:</b> {'; '.join(excluded)}", self.body_style))

        # 1.4 Allocation
        elements.append(Paragraph("1.4 Allocation Procedures", self.section_style))
        alloc = gs.get("allocation", {})
        if isinstance(alloc, dict):
            elements.append(Paragraph(f"<b>Principle:</b> {alloc.get('principle', 'ALLOCATION')}", self.body_style))
            elements.append(Paragraph(f"<b>Method:</b> {alloc.get('method', 'MASS')}", self.body_style))
            elements.append(Paragraph(f"<b>Recycling Method:</b> {alloc.get('recyclingMethod', 'CUTOFF')}", self.body_style))

        # 1.5 LCIA Methodology
        elements.append(Paragraph("1.5 Impact Assessment Methodology", self.section_style))
        lcia = gs.get("lcia", {})
        if isinstance(lcia, dict):
            elements.append(Paragraph(f"<b>Methodology:</b> {lcia.get('methodology', 'EF_3_1')}", self.body_style))
            cats = lcia.get("categories", [])
            if cats:
                elements.append(Paragraph(f"<b>Selected Categories:</b> {', '.join(cats)}", self.body_style))

        # 1.6 Data Quality
        elements.append(Paragraph("1.6 Data Quality Requirements", self.section_style))
        dq = gs.get("dataQuality", {})
        if isinstance(dq, dict):
            elements.append(Paragraph(f"<b>Temporal Coverage:</b> {dq.get('timeframe', 'Not specified')}", self.body_style))
            elements.append(Paragraph(f"<b>Geographical Coverage:</b> {dq.get('geography', 'Not specified')}", self.body_style))
            elements.append(Paragraph(f"<b>Technological Coverage:</b> {dq.get('technology', 'Not specified')}", self.body_style))

        # 1.7 Review
        elements.append(Paragraph("1.7 Type of Critical Review", self.section_style))
        review = gs.get("review", {})
        if isinstance(review, dict):
            elements.append(Paragraph(f"<b>Review Type:</b> {review.get('type', 'INTERNAL')}", self.body_style))
        else:
            elements.append(Paragraph(f"<b>Review Type:</b> {review}", self.body_style))

        # Build
        doc.build(elements)
        buf.seek(0)
        return buf

    # ═══════════════════════════════════════════════════════════════════════
    # 02. LCI EXCEL
    # ═══════════════════════════════════════════════════════════════════════

    def _generate_lci_excel(self, nodes: List[Dict], edges: List[Dict]) -> io.BytesIO:
        try:
            import openpyxl
        except ImportError:
            # Fallback: use a simple CSV-in-buffer approach
            return self._generate_lci_csv_fallback(nodes, edges)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "LCI Process Tree"

        # Header
        headers = ["Process Name", "Module", "Location", "Flow Name", "Amount", "Unit", "Direction", "Category", "Formula", "Source"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill("solid", fgColor="1A5F7A")

        row = 2
        for n in nodes:
            data = n.get("data", {})
            proc_name = data.get("processName", data.get("label", "Unnamed"))
            module = data.get("module", "A1-A3")
            location = data.get("location", "GLO")
            if isinstance(location, dict):
                location = location.get("value", "GLO")

            exchanges = data.get("exchanges", [])
            inputs = data.get("inputs", [])
            outputs = data.get("outputs", [])

            all_flows = []
            for ex in exchanges:
                all_flows.append({
                    "name": ex.get("flow_name", ex.get("name", "")),
                    "amount": ex.get("amount", 0),
                    "unit": ex.get("unit", "kg"),
                    "direction": "Input" if ex.get("is_input", True) else "Output",
                    "category": ex.get("category", ""),
                    "formula": ex.get("formula", ""),
                })
            for inp in inputs:
                all_flows.append({
                    "name": inp.get("name", ""), "amount": inp.get("amount", 0),
                    "unit": inp.get("unit", "kg"), "direction": "Input",
                    "category": inp.get("category", ""), "formula": inp.get("formula", ""),
                })
            for out in outputs:
                all_flows.append({
                    "name": out.get("name", ""), "amount": out.get("amount", 0),
                    "unit": out.get("unit", "kg"), "direction": "Output",
                    "category": out.get("category", ""), "formula": out.get("formula", ""),
                })

            if not all_flows:
                all_flows = [{"name": "(no exchanges)", "amount": 0, "unit": "-", "direction": "-", "category": "", "formula": ""}]

            source = "Database" if data.get("processId") or data.get("is_library") else "Manual"

            for flow in all_flows:
                ws.cell(row=row, column=1, value=proc_name)
                ws.cell(row=row, column=2, value=module)
                ws.cell(row=row, column=3, value=str(location))
                ws.cell(row=row, column=4, value=flow["name"])
                ws.cell(row=row, column=5, value=float(flow.get("amount", 0)))
                ws.cell(row=row, column=6, value=flow["unit"])
                ws.cell(row=row, column=7, value=flow["direction"])
                ws.cell(row=row, column=8, value=flow.get("category", ""))
                ws.cell(row=row, column=9, value=flow.get("formula", ""))
                ws.cell(row=row, column=10, value=source)
                row += 1

        # Edge sheet
        ws_edges = wb.create_sheet("Process Links")
        edge_headers = ["Edge ID", "Source Node", "Target Node"]
        for col, h in enumerate(edge_headers, 1):
            cell = ws_edges.cell(row=1, column=col, value=h)
            cell.font = openpyxl.styles.Font(bold=True)
        for i, e in enumerate(edges, 2):
            ws_edges.cell(row=i, column=1, value=e.get("id", ""))
            ws_edges.cell(row=i, column=2, value=str(e.get("source", "")))
            ws_edges.cell(row=i, column=3, value=str(e.get("target", "")))

        # Auto-width
        for ws_sheet in [ws, ws_edges]:
            for col in ws_sheet.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws_sheet.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def _generate_lci_csv_fallback(self, nodes: List[Dict], edges: List[Dict]) -> io.BytesIO:
        """Fallback if openpyxl is not installed."""
        import csv
        buf = io.BytesIO()
        text_buf = io.TextIOWrapper(buf, encoding='utf-8', newline='')
        writer = csv.writer(text_buf)
        writer.writerow(["Process", "Module", "Flow", "Amount", "Unit", "Direction"])
        for n in nodes:
            data = n.get("data", {})
            for ex in data.get("exchanges", []):
                writer.writerow([
                    data.get("processName", ""), data.get("module", ""),
                    ex.get("flow_name", ""), ex.get("amount", 0),
                    ex.get("unit", "kg"), "Input" if ex.get("is_input") else "Output"
                ])
        text_buf.flush()
        text_buf.detach()
        buf.seek(0)
        return buf

    # ═══════════════════════════════════════════════════════════════════════
    # 03. LCIA RESULTS EXCEL
    # ═══════════════════════════════════════════════════════════════════════

    def _generate_lcia_excel(self, results: Dict) -> io.BytesIO:
        try:
            import openpyxl
        except ImportError:
            buf = io.BytesIO()
            buf.write(json.dumps(results.get("impacts", {}), indent=2).encode())
            buf.seek(0)
            return buf

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "LCIA Results"

        # Summary Sheet
        headers = ["Impact Category", "Result", "Unit", "Source"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill("solid", fgColor="1A5F7A")

        impacts = results.get("impacts", {})
        for row, cat in enumerate(IMPACT_CATEGORIES, 2):
            val = impacts.get(cat, 0.0)
            ws.cell(row=row, column=1, value=cat.replace("_", " ").title())
            ws.cell(row=row, column=2, value=float(val))
            ws.cell(row=row, column=2).number_format = '0.0000E+00'
            ws.cell(row=row, column=3, value=UNIT_MAP.get(cat, "pts"))
            ws.cell(row=row, column=4, value="Deterministic" if not results.get("uncertainty") else "Monte Carlo Mean")

        # Node Breakdown Sheet
        breakdown = results.get("node_breakdown", {})
        if breakdown:
            ws2 = wb.create_sheet("Node Breakdown")
            bd_headers = ["Node ID", "Process Name", "Module"] + [c.upper() for c in IMPACT_CATEGORIES[:6]] + ["..."]
            for col, h in enumerate(bd_headers, 1):
                cell = ws2.cell(row=1, column=col, value=h)
                cell.font = openpyxl.styles.Font(bold=True)

            for row, (nid, info) in enumerate(breakdown.items(), 2):
                ws2.cell(row=row, column=1, value=str(nid))
                ws2.cell(row=row, column=2, value=info.get("name", ""))
                ws2.cell(row=row, column=3, value=info.get("module", ""))
                node_impacts = info.get("impacts", {})
                for col_off, cat in enumerate(IMPACT_CATEGORIES[:6]):
                    ws2.cell(row=row, column=4 + col_off, value=float(node_impacts.get(cat, 0)))

        # Uncertainty Sheet (if Monte Carlo)
        uncertainty = results.get("uncertainty")
        if uncertainty:
            ws3 = wb.create_sheet("Uncertainty Analysis")
            unc_headers = ["Category", "Mean", "Median", "Std Dev", "P5", "P95"]
            for col, h in enumerate(unc_headers, 1):
                ws3.cell(row=1, column=col, value=h)
            for row, cat in enumerate(IMPACT_CATEGORIES, 2):
                unc = uncertainty.get(cat, {})
                ws3.cell(row=row, column=1, value=cat.replace("_", " ").title())
                ws3.cell(row=row, column=2, value=unc.get("mean", 0))
                ws3.cell(row=row, column=3, value=unc.get("median", 0))
                ws3.cell(row=row, column=4, value=unc.get("std", 0))
                ws3.cell(row=row, column=5, value=unc.get("p5", 0))
                ws3.cell(row=row, column=6, value=unc.get("p95", 0))

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # ═══════════════════════════════════════════════════════════════════════
    # 04. LCIA REPORT PDF
    # ═══════════════════════════════════════════════════════════════════════

    def _generate_lcia_report_pdf(self, results: Dict, gs: Dict, framework: str) -> io.BytesIO:
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=50, bottomMargin=50)
        elements = []

        elements.append(Paragraph("LIFE CYCLE IMPACT ASSESSMENT REPORT", self.header_style))
        elements.append(Paragraph(
            f"<i>{framework.upper()} | {gs.get('projectTitle', 'LCA Study')} | "
            f"{datetime.datetime.now().strftime('%Y-%m-%d')}</i>",
            self.small_style
        ))

        # Summary table
        elements.append(Paragraph("Impact Assessment Results", self.section_style))
        impacts = results.get("impacts", {})
        table_data = [["Impact Category", "Result", "Unit"]]
        for cat in IMPACT_CATEGORIES:
            val = impacts.get(cat, 0.0)
            table_data.append([
                cat.replace("_", " ").title(),
                f"{val:.4e}",
                UNIT_MAP.get(cat, "pts")
            ])

        t = Table(table_data, colWidths=[190, 130, 130])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.toColor("#1a5f7a")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 20))

        # Node contribution
        breakdown = results.get("node_breakdown", {})
        if breakdown:
            elements.append(Paragraph("Contribution Analysis", self.section_style))
            gwp_total = impacts.get("gwp_climate_change", 1.0)
            contrib_data = [["Process", "GWP (kg CO2-eq)", "Contribution %"]]
            sorted_nodes = sorted(
                breakdown.items(),
                key=lambda x: x[1].get("impacts", {}).get("gwp_climate_change", 0),
                reverse=True
            )
            for nid, info in sorted_nodes[:10]:
                gwp = info.get("impacts", {}).get("gwp_climate_change", 0)
                pct = (gwp / gwp_total * 100) if gwp_total > 0 else 0
                contrib_data.append([
                    info.get("name", str(nid))[:40],
                    f"{gwp:.4e}",
                    f"{pct:.1f}%"
                ])

            ct = Table(contrib_data, colWidths=[200, 120, 100])
            ct.setStyle(TableStyle([
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
                ('BACKGROUND', (0, 0), (-1, 0), colors.toColor("#f0f4f8")),
                ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
            ]))
            elements.append(ct)

        doc.build(elements)
        buf.seek(0)
        return buf

    # ═══════════════════════════════════════════════════════════════════════
    # 05. INTERPRETATION REPORT PDF
    # ═══════════════════════════════════════════════════════════════════════

    def _generate_interpretation_pdf(
        self, results: Dict, sensitivity: Optional[Dict], gs: Dict
    ) -> io.BytesIO:
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=50, bottomMargin=50)
        elements = []

        elements.append(Paragraph("INTERPRETATION REPORT", self.header_style))

        # Deep interpretation if available
        interp = results.get("deep_interpretation", {})

        # 5.1 Narrative
        elements.append(Paragraph("5.1 Significant Issues", self.section_style))
        narrative = interp.get("narrative", f"Total Climate Impact: {results.get('gwp', 0):.4e} kg CO2-eq.")
        elements.append(Paragraph(narrative, self.body_style))

        # 5.2 Hotspots
        elements.append(Paragraph("5.2 Hotspot Analysis", self.section_style))
        hotspots = interp.get("hotspots", [])
        if hotspots:
            hs_data = [["Rank", "Process", "Contribution %"]]
            for i, hs in enumerate(hotspots, 1):
                hs_data.append([str(i), hs.get("name", "N/A"), f"{hs.get('contribution', 0):.1f}%"])
            ht = Table(hs_data, colWidths=[50, 250, 100])
            ht.setStyle(TableStyle([
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
            ]))
            elements.append(ht)
        else:
            elements.append(Paragraph("No significant hotspots identified.", self.body_style))

        # 5.3 Sensitivity
        elements.append(Paragraph("5.3 Sensitivity Analysis", self.section_style))
        if sensitivity:
            elements.append(Paragraph(
                f"Node: {sensitivity.get('node_name', 'N/A')}<br/>"
                f"Baseline GWP: {sensitivity.get('baseline', 0):.4e} kg CO2-eq<br/>"
                f"Low Scenario (−10%): {sensitivity.get('low_scenario', 0):.4e} ({sensitivity.get('variance_low_pct', 0):+.2f}%)<br/>"
                f"High Scenario (+10%): {sensitivity.get('high_scenario', 0):.4e} ({sensitivity.get('variance_high_pct', 0):+.2f}%)",
                self.body_style
            ))
        else:
            elements.append(Paragraph(
                "No sensitivity analysis was performed for this study. "
                "Per ISO 14044, sensitivity analysis should be conducted on key assumptions.",
                self.body_style
            ))

        # 5.4 Recommendations
        elements.append(Paragraph("5.4 Recommendations", self.section_style))
        advice = interp.get("advice", "No critical optimizations detected.")
        elements.append(Paragraph(advice, self.body_style))

        # 5.5 Benchmarks
        benchmarks = interp.get("benchmarks", {})
        if benchmarks:
            elements.append(Paragraph("5.5 Benchmark Comparison", self.section_style))
            bm_data = [["Category", "Study Result", "World Avg/capita", "Status"]]
            for cat, bm in benchmarks.items():
                bm_data.append([
                    cat.replace("_", " ").title()[:30],
                    f"{bm.get('value', 0):.3e}",
                    f"{bm.get('world_avg', 0):.3e}",
                    bm.get("status", "N/A")
                ])
            bt = Table(bm_data, colWidths=[140, 100, 100, 80])
            bt.setStyle(TableStyle([
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
                ('BACKGROUND', (0, 0), (-1, 0), colors.toColor("#f0f4f8")),
            ]))
            elements.append(bt)

        doc.build(elements)
        buf.seek(0)
        return buf

    # ═══════════════════════════════════════════════════════════════════════
    # 06. DATA QUALITY ASSESSMENT PDF
    # ═══════════════════════════════════════════════════════════════════════

    def _generate_data_quality_pdf(self, nodes: List[Dict], gs: Dict) -> io.BytesIO:
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=50, bottomMargin=50)
        elements = []

        elements.append(Paragraph("DATA QUALITY ASSESSMENT", self.header_style))

        # DQ Requirements from Goal & Scope
        elements.append(Paragraph("6.1 Data Quality Requirements", self.section_style))
        dq = gs.get("dataQuality", {})
        if isinstance(dq, dict):
            elements.append(Paragraph(f"Temporal: {dq.get('timeframe', 'Not specified')}", self.body_style))
            elements.append(Paragraph(f"Geographical: {dq.get('geography', 'Not specified')}", self.body_style))
            elements.append(Paragraph(f"Technological: {dq.get('technology', 'Not specified')}", self.body_style))

        # Pedigree Matrix per Node
        elements.append(Paragraph("6.2 Process-Level Pedigree Assessment", self.section_style))

        ped_data = [["Process", "Reliability", "Completeness", "Temporal", "Geographical", "Technological", "Source"]]
        for n in nodes:
            data = n.get("data", {})
            label = data.get("processName", data.get("label", "Unnamed"))[:30]
            pedigree = data.get("pedigree", data.get("metadata", {}).get("pedigree_matrix", {}))

            if isinstance(pedigree, dict):
                r = pedigree.get("reliability", 3)
                c = pedigree.get("completeness", 3)
                t = pedigree.get("temporal_correlation", pedigree.get("temporal", 3))
                g = pedigree.get("geographical_correlation", pedigree.get("geographical", 3))
                tech = pedigree.get("technological_correlation", pedigree.get("technological", 3))
            else:
                r = c = t = g = tech = 3  # Default moderate quality

            source = "Library" if data.get("processId") or data.get("is_library") else "Manual"
            ped_data.append([label, str(r), str(c), str(t), str(g), str(tech), source])

        if len(ped_data) > 1:
            pt = Table(ped_data, colWidths=[100, 55, 65, 55, 65, 65, 55])
            pt.setStyle(TableStyle([
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
                ('BACKGROUND', (0, 0), (-1, 0), colors.toColor("#f0f4f8")),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ]))
            elements.append(pt)
        else:
            elements.append(Paragraph("No process nodes available for quality assessment.", self.body_style))

        # Score interpretation
        elements.append(Spacer(1, 20))
        elements.append(Paragraph("6.3 Score Interpretation (ISO 14044 / ecoinvent Pedigree)", self.section_style))
        legend = [
            ["Score", "Meaning"],
            ["1", "Verified data from the specific process"],
            ["2", "Non-verified data from the same process or verified from similar"],
            ["3", "Non-verified data from similar processes or estimated"],
            ["4", "Estimated data with limited verification"],
            ["5", "Non-qualified estimate, unknown origin"],
        ]
        lt = Table(legend, colWidths=[50, 400])
        lt.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
        ]))
        elements.append(lt)

        doc.build(elements)
        buf.seek(0)
        return buf

    # ═══════════════════════════════════════════════════════════════════════
    # 07. CRITICAL REVIEW CHECKLIST PDF
    # ═══════════════════════════════════════════════════════════════════════

    def _generate_review_checklist_pdf(
        self, validation: Optional[Dict], gs: Dict, framework: str
    ) -> io.BytesIO:
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=50, bottomMargin=50)
        elements = []

        elements.append(Paragraph("CRITICAL REVIEW CHECKLIST", self.header_style))
        elements.append(Paragraph(
            f"<i>Per ISO 14044 Clause 6 | Framework: {framework.upper()}</i>",
            self.small_style
        ))

        if validation and "checks" in validation:
            elements.append(Paragraph("Automated Compliance Checks", self.section_style))

            check_data = [["#", "Check", "Standard", "Result", "Severity"]]
            for i, check in enumerate(validation["checks"], 1):
                status = "✓ PASS" if check["passed"] else "✗ FAIL"
                check_data.append([
                    str(i),
                    check["name"][:40],
                    check["standard"],
                    status,
                    check["severity"]
                ])

            ct = Table(check_data, colWidths=[25, 180, 70, 65, 60])
            ct.setStyle(TableStyle([
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
                ('BACKGROUND', (0, 0), (-1, 0), colors.toColor("#1a5f7a")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ]))
            elements.append(ct)

            # Summary
            elements.append(Spacer(1, 20))
            level = validation.get("compliance_level", "NONE")
            total = validation.get("total_checks", 0)
            critical = validation.get("critical_failures", 0)
            warnings = validation.get("warnings", 0)
            passed = total - critical - warnings

            elements.append(Paragraph(
                f"<b>Compliance Level:</b> {level} | "
                f"<b>Passed:</b> {passed}/{total} | "
                f"<b>Critical Failures:</b> {critical} | "
                f"<b>Warnings:</b> {warnings}",
                self.body_style
            ))

            # Failed check details
            failed = [c for c in validation["checks"] if not c["passed"]]
            if failed:
                elements.append(Paragraph("Required Actions", self.section_style))
                for c in failed:
                    elements.append(Paragraph(
                        f"<b>[{c['severity']}] {c['name']}:</b> {c['message']}",
                        self.body_style
                    ))
                    if c.get("recommendation"):
                        elements.append(Paragraph(
                            f"<i>→ {c['recommendation']}</i>",
                            self.small_style
                        ))
        else:
            elements.append(Paragraph(
                "No automated validation was performed. Run the study validator before generating the review checklist.",
                self.body_style
            ))

        # Reviewer Signature Block
        elements.append(Spacer(1, 40))
        elements.append(Paragraph("Reviewer Declaration", self.section_style))
        sig_data = [
            ["Reviewer Name:", "____________________________", "Date:", "____________________________"],
            ["Organization:", "____________________________", "Qualification:", "____________________________"],
            ["Signature:", "____________________________", "", ""],
        ]
        st = Table(sig_data, colWidths=[80, 160, 80, 160])
        st.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 15),
        ]))
        elements.append(st)

        doc.build(elements)
        buf.seek(0)
        return buf

    # ═══════════════════════════════════════════════════════════════════════
    # METADATA JSON
    # ═══════════════════════════════════════════════════════════════════════

    def _generate_metadata(self, gs: Dict, results: Dict, framework: str) -> Dict:
        fu = gs.get("functionalUnit", {})
        return {
            "schema": "triya-lca-study-v1",
            "generated": datetime.datetime.now().isoformat(),
            "framework": framework,
            "project": gs.get("projectTitle", "Untitled"),
            "functionalUnit": {
                "description": fu.get("description", "") if isinstance(fu, dict) else str(fu),
                "magnitude": fu.get("magnitude", 1.0) if isinstance(fu, dict) else 1.0,
                "unit": fu.get("unit", "unit") if isinstance(fu, dict) else "unit",
            },
            "systemBoundary": gs.get("systemBoundary", {}).get("scope", "CRADLE_TO_GATE") if isinstance(gs.get("systemBoundary"), dict) else "CRADLE_TO_GATE",
            "lciaMethodology": gs.get("lcia", {}).get("methodology", "EF_3_1") if isinstance(gs.get("lcia"), dict) else "EF_3_1",
            "totalImpacts": {
                cat: results.get("impacts", {}).get(cat, 0.0) for cat in IMPACT_CATEGORIES
            },
            "gwpTotal": results.get("gwp", results.get("impacts", {}).get("gwp_climate_change", 0.0)),
            "isAiPredicted": results.get("is_ai_predicted", False),
            "iterations": results.get("iterations", 1),
            "generator": "Triya.io StudyPackageGenerator v1.0"
        }
